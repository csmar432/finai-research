"""End-to-end coverage for the deterministic national workbook builder."""

from __future__ import annotations

from openpyxl import load_workbook

from scripts import generate_national_excel as mod


def test_main_builds_all_sheets_from_versioned_dataset(tmp_path, monkeypatch, capsys):
    output = tmp_path / "national.xlsx"
    monkeypatch.setattr(mod, "OUT_FILE", output)
    # Use the checked-in dataset: this exercises the real formatting, ranking,
    # time-series, verification, and source-tracking paths without network I/O.
    mod.main()
    assert output.exists() and output.stat().st_size > 10_000
    wb = load_workbook(output, read_only=True)
    assert wb.sheetnames == [
        "1-数据总览", "2-GDP对比", "3-RD对比", "4-科技企业对比",
        "5-技术转化对比", "6-新兴产业对比", "7-高校与人才",
        "8-湖北时间序列", "9-数据质量报告",
    ]
    assert wb["1-数据总览"]["A1"].value.startswith("全国各省")
    assert wb["9-数据质量报告"]["A1"].value.startswith("数据质量报告")
    assert "9 sheets" in capsys.readouterr().out

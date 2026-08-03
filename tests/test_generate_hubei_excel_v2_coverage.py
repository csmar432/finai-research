"""Exercise every deterministic sheet in the Hubei workbook generator."""

from __future__ import annotations

from openpyxl import load_workbook

from scripts import generate_hubei_excel_v2 as mod


def test_main_writes_nine_complete_sheets(tmp_path, monkeypatch, capsys):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "data").mkdir()
    mod.main()
    output = tmp_path / "data" / "湖北省科技创新数据_2026_v2.xlsx"
    assert output.exists() and output.stat().st_size > 10_000
    wb = load_workbook(output, read_only=True)
    assert wb.sheetnames == mod.TABS
    assert wb[mod.TABS[0]]["A1"].value.startswith("湖北省科技创新")
    assert "9 sheets" in capsys.readouterr().out

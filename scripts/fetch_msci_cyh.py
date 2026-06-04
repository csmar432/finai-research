#!/usr/bin/env python3
"""批量获取陈宇浩缺失的MSCI ESG评级"""
import json
import os
import ssl
import time
import urllib.request

import openpyxl

ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Referer': 'https://finance.sina.com.cn/esg/grade.shtml',
}

EXCEL_PATH = str(Path.home() / 'Desktop' / '2026MSCI级别6.1(1).xlsx')
OUTPUT_PATH = str(Path.home() / 'Desktop' / '2026MSCI级别6.1(1)_已填充.xlsx')
CACHE_FILE = str(Path.home() / 'Desktop' / '论文-研报工作流/data/msci_esg_ratings_cyh.json')

print("1. 读取Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['2025年度上市公司名单']

cyh_stocks = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if row[6] == '陈宇浩':
        ts_code = row[1]
        if ts_code and ts_code.endswith('.SH'):
            symbol = 'sh' + ts_code.replace('.SH', '')
        elif ts_code and ts_code.endswith('.SZ'):
            symbol = 'sz' + ts_code.replace('.SZ', '')
        else:
            symbol = None
        cyh_stocks.append({
            'row_idx': row_idx,
            'name': row[0],
            'ts_code': ts_code,
            'hk_code': row[2],
            'short': row[3],
            'symbol': symbol,
            'rating': row[4],
        })

unfilled = [s for s in cyh_stocks if s['rating'] is None]
print(f"   陈宇浩总数: {len(cyh_stocks)}")
print(f"   已填充: {sum(1 for s in cyh_stocks if s['rating'] is not None)}")
print(f"   未填充: {len(unfilled)}")

# 加载缓存
cache = {}
if os.path.exists(CACHE_FILE):
    with open(CACHE_FILE, encoding='utf-8') as f:
        cache = json.load(f)
    print(f"   缓存已有: {len(cache)} 条")

# 额外检查许哲逸的缓存
XU_CACHE = str(Path.home() / 'Desktop' / '论文-研报工作流/data/msci_esg_ratings_xu.json')
if os.path.exists(XU_CACHE):
    with open(XU_CACHE, encoding='utf-8') as f:
        xu_cache = json.load(f)
    print(f"   许哲逸缓存: {len(xu_cache)} 条 (可复用)")

to_fetch = []
for s in unfilled:
    if s['symbol']:
        sym_up = s['symbol'].upper()
        if sym_up not in cache:
            # 也检查许哲逸缓存
            if XU_CACHE and os.path.exists(XU_CACHE) and sym_up in xu_cache:
                cache[sym_up] = xu_cache[sym_up]
            else:
                to_fetch.append(s)

print(f"   需新获取: {len(to_fetch)} 只")

def get_msci(symbol):
    url = f'https://global.finance.sina.com.cn/api/openapi.php/EsgService.getEsgStockInfo?symbol={symbol}'
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=15, context=ctx) as response:
            data = json.loads(response.read().decode('utf-8'))
            for info in data.get('result', {}).get('data', {}).get('info', []):
                if info.get('agency_name') == 'MSCI':
                    return info.get('esg_score'), info.get('esg_dt')
    except Exception:
        pass
    return None, None

if to_fetch:
    print("\n2. 开始获取...")
    fetched = 0
    start = time.time()
    for i, s in enumerate(to_fetch):
        sym = s['symbol']
        rating, date = get_msci(sym)
        cache[sym.upper()] = {'msci_rating': rating, 'msci_date': date}
        if rating:
            fetched += 1
        if (i+1) % 25 == 0:
            elapsed = time.time() - start
            rate = elapsed / (i+1)
            eta = rate * (len(to_fetch) - i - 1) if (i+1) > 0 else 0
            print(f"   {i+1}/{len(to_fetch)} | 成功: {fetched} | ETA: {eta/60:.1f}min")
        time.sleep(0.3)

    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)
    print(f"   缓存已保存, 新增成功: {fetched}")

# 回填Excel
print("\n3. 回填Excel...")
filled = 0
no_rating = 0

for s in cyh_stocks:
    row_idx = s['row_idx']
    if s['rating'] is not None:
        continue
    if not s['symbol']:
        no_rating += 1
        continue
    sym_upper = s['symbol'].upper()
    if sym_upper in cache:
        result = cache[sym_upper]
        rating = result.get('msci_rating')
        date = result.get('msci_date')
        if rating and rating != '-':
            ws.cell(row=row_idx, column=5, value=rating)
            ws.cell(row=row_idx, column=6, value=date)
            filled += 1
        else:
            no_rating += 1
    else:
        no_rating += 1

wb.save(OUTPUT_PATH)
print(f"   保存: {OUTPUT_PATH}")
total_filled = sum(1 for s in cyh_stocks if s['rating'] is not None or (s['symbol'] and cache.get(s['symbol'].upper(),{}).get('msci_rating')))
print(f"   本次新增填充: {filled}")
print(f"   陈宇浩总计已填充: {total_filled}/{len(cyh_stocks)}")

# 分布
dist = {}
for s in cyh_stocks:
    if s['symbol']:
        sym_up = s['symbol'].upper()
        r = s['rating'] if s['rating'] else cache.get(sym_up, {}).get('msci_rating')
        if r and r != '-':
            dist[r] = dist.get(r, 0) + 1

print("\n=== 评级分布 ===")
for r in sorted(dist.keys()):
    print(f"  {r}: {dist[r]}")
print("Done!")

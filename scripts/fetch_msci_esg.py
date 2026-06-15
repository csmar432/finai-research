#!/usr/bin/env python3
"""
fetch_msci_esg.py — 批量获取股票的 MSCI ESG 评级 (新浪财经 API)

通用工具：根据 Excel 中指定负责人列, 批量填充 MSCI ESG 评级。

使用:
  python scripts/fetch_msci_esg.py \\
    --person "<your_name>" \\
    --excel /path/to/input.xlsx \\
    --output /path/to/output.xlsx

环境变量:
  EXCEL_PATH   — 输入 Excel 文件路径
  OUTPUT_PATH  — 输出 Excel 文件路径
  CACHE_DIR    — 缓存目录 (默认: data/)
  PERSON_NAME  — Excel 中负责人的列值 (默认: "your_name")
"""
import argparse
import concurrent.futures
import json
import os
import ssl
import time
import urllib.request
from pathlib import Path

import openpyxl

# ── 路径配置 ────────────────────────────────────────────────────
_PROJECT_ROOT = Path(__file__).parent.parent
_CACHE_DIR = Path(os.environ.get("CACHE_DIR", str(_PROJECT_ROOT / "data")))
DEFAULT_EXCEL_PATH = str(_PROJECT_ROOT / "data" / "msci_input.xlsx")
DEFAULT_OUTPUT_PATH = str(_PROJECT_ROOT / "data" / "msci_output.xlsx")
DEFAULT_PERSON = "your_name"


def _get_paths(args_namespace) -> tuple[str, str]:
    """从参数或环境变量获取路径, 文件不存在时给出清晰错误。"""
    excel = args_namespace.excel if args_namespace.excel else DEFAULT_EXCEL_PATH
    output = args_namespace.output if args_namespace.output else DEFAULT_OUTPUT_PATH
    if not os.path.exists(excel):
        raise FileNotFoundError(
            f"输入 Excel 不存在: {excel}\n"
            f"请通过 --excel 参数指定文件路径, 或设置 EXCEL_PATH 环境变量。"
        )
    return excel, output


# === SSL 上下文 ===
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Referer': 'https://finance.sina.com.cn/esg/grade.shtml',
}


def normalize_symbol(ts_code: str, hk_code: str) -> str:
    """将 tushare 格式转换为新浪 API 格式"""
    if ts_code is None or ts_code == '':
        if hk_code:
            return hk_code.upper().replace('.HK', 'hk').replace('HK', 'hk')
        return None

    ts = ts_code.upper()
    if ts.endswith('.SH'):
        return 'sh' + ts.replace('.SH', '')
    elif ts.endswith('.SZ'):
        return 'sz' + ts.replace('.SZ', '')
    elif ts.endswith('.BJ'):
        return None
    return None


def get_esg_info(symbol: str, max_retries: int = 2) -> dict:
    """获取单只股票的 ESG 数据, 返回 MSCI 评级和日期"""
    if not symbol:
        return {'symbol': symbol, 'msci_rating': None, 'msci_date': None}

    url = f'https://global.finance.sina.com.cn/api/openapi.php/EsgService.getEsgStockInfo?symbol={symbol}'
    for attempt in range(max_retries):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=15, context=ctx) as response:
                try:
                    data = json.loads(response.read().decode('utf-8'))
                except json.JSONDecodeError:
                    return {'symbol': symbol, 'msci_rating': None, 'msci_date': None}
                info_list = data.get('result', {}).get('data', {}).get('info', [])
                for info in info_list:
                    if info.get('agency_name') == 'MSCI':
                        return {
                            'symbol': symbol.upper(),
                            'msci_rating': info.get('esg_score'),
                            'msci_date': info.get('esg_dt'),
                        }
                return {'symbol': symbol.upper(), 'msci_rating': None, 'msci_date': None}
        except Exception:
            if attempt < max_retries - 1:
                time.sleep(1)
            return {'symbol': symbol.upper(), 'msci_rating': None, 'msci_date': None}
    return {'symbol': symbol.upper(), 'msci_rating': None, 'msci_date': None}


def main():
    parser = argparse.ArgumentParser(description="批量获取股票的 MSCI ESG 评级")
    parser.add_argument(
        '--person',
        default=os.environ.get('PERSON_NAME', DEFAULT_PERSON),
        help='Excel 中负责人的列值 (环境变量: PERSON_NAME)',
    )
    parser.add_argument('--excel', help='输入 Excel 路径 (环境变量: EXCEL_PATH)')
    parser.add_argument('--output', help='输出 Excel 路径 (环境变量: OUTPUT_PATH)')
    parser.add_argument(
        '--sheet',
        default='2025年度上市公司名单',
        help='Excel sheet 名称 (默认: 2025年度上市公司名单)',
    )
    args = parser.parse_args()

    excel, output = _get_paths(args)
    person = args.person
    cache_key = person.replace(' ', '_')
    cache_path = str(_CACHE_DIR / f"msci_esg_ratings_{cache_key}.json")

    # 1. 读取 Excel
    print("1. 读取 Excel 文件...")
    wb = openpyxl.load_workbook(excel)
    ws = wb[args.sheet]

    target_stocks = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) > 6 and row[6] == person:
            ts_code = row[1]
            hk_code = row[2]
            symbol = normalize_symbol(ts_code, hk_code)
            target_stocks.append({
                'row_idx': row_idx,
                'name': row[0],
                'ts_code': ts_code,
                'hk_code': hk_code,
                'short_name': row[3],
                'symbol': symbol,
            })

    print(f"   [{person}] 负责股票数: {len(target_stocks)}")
    stocks_with_symbol = [s for s in target_stocks if s['symbol']]
    print(f"   有 API 符号的股票: {len(stocks_with_symbol)}")
    stocks_no_symbol = [s for s in target_stocks if not s['symbol']]
    if stocks_no_symbol:
        print(f"   无法获取的股票 (无符号): {len(stocks_no_symbol)}")

    # 2. 加载已有缓存
    cache = {}
    if os.path.exists(cache_path):
        with open(cache_path, encoding='utf-8') as f:
            cache = json.load(f)
        print(f"   加载缓存: {len(cache)} 条")

    # 3. 确定需要新获取的股票
    to_fetch = []
    for s in stocks_with_symbol:
        sym_upper = s['symbol'].upper()
        if sym_upper not in cache:
            to_fetch.append(s)

    print(f"   需要新获取: {len(to_fetch)} 只")
    print(f"   已命中缓存: {len(stocks_with_symbol) - len(to_fetch)} 只")

    if to_fetch:
        print("\n2. 开始获取 MSCI 数据 (并发 5 个 worker)...")
        fetched = 0
        failed = 0
        start_time = time.time()

        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            futures = {executor.submit(get_esg_info, s['symbol']): s for s in to_fetch}
            done = 0
            for future in concurrent.futures.as_completed(futures):
                stock = futures[future]
                result = future.result()
                sym_upper = stock['symbol'].upper()
                cache[sym_upper] = result
                done += 1

                if result.get('msci_rating'):
                    fetched += 1
                else:
                    failed += 1

                if done % 50 == 0:
                    elapsed = time.time() - start_time
                    rate = done / elapsed
                    remaining = len(to_fetch) - done
                    eta = remaining / rate if rate > 0 else 0
                    print(f"   进度: {done}/{len(to_fetch)} ({100*done/len(to_fetch):.1f}%) | "
                          f"成功: {fetched} | 失败: {failed} | "
                          f"ETA: {eta/60:.1f}min")

                time.sleep(0.2)

        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
        print(f"\n   缓存已保存: {cache_path}")
        print(f"   本轮新增成功: {fetched}")
        print(f"   本轮失败: {failed}")

    # 4. 回填 Excel
    print("\n3. 回填 Excel 数据...")
    filled_count = 0
    no_rating_count = 0
    no_symbol_count = 0

    for s in target_stocks:
        row_idx = s['row_idx']
        if not s['symbol']:
            no_symbol_count += 1
            continue

        sym_upper = s['symbol'].upper()
        if sym_upper in cache:
            result = cache[sym_upper]
            rating = result.get('msci_rating')
            date = result.get('msci_date')
            if rating and rating != '-':
                ws.cell(row=row_idx, column=5, value=rating)
                ws.cell(row=row_idx, column=6, value=date)
                filled_count += 1
            else:
                no_rating_count += 1
        else:
            no_rating_count += 1

    print(f"   成功填充 MSCI 评级: {filled_count}")
    print(f"   无 MSCI 评级: {no_rating_count}")
    print(f"   无 API 符号 (北交所等): {no_symbol_count}")

    # 5. 保存
    print("\n4. 保存文件...")
    wb.save(output)
    print(f"   已保存: {output}")

    # 6. 统计总览
    print("\n=== 汇总 ===")
    total = len(target_stocks)
    print(f"[{person}] 负责总数: {total}")
    if total > 0:
        print(f"成功获取评级: {filled_count} ({100*filled_count/total:.1f}%)")
        print(f"无 MSCI 评级: {no_rating_count} ({100*no_rating_count/total:.1f}%)")
        print(f"无 API 符号: {no_symbol_count} ({100*no_symbol_count/total:.1f}%)")

    rating_dist = {}
    for s in target_stocks:
        if s['symbol']:
            sym_upper = s['symbol'].upper()
            if sym_upper in cache:
                r = cache[sym_upper].get('msci_rating', '-')
                if r and r != '-':
                    rating_dist[r] = rating_dist.get(r, 0) + 1
    print("\nMSCI 评级分布:")
    for r in sorted(rating_dist.keys()):
        print(f"  {r}: {rating_dist[r]}")


if __name__ == '__main__':
    main()

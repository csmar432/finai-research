#!/usr/bin/env python3
"""
fetch_msci_esg_v2.py — 批量获取 MSCI ESG 评级 (顺序版, 已被 fetch_msci_esg.py 替代)

⚠️  DEPRECATED — 此脚本已被 fetch_msci_esg.py 替代
   保留此文件仅用于历史参考, 建议使用 fetch_msci_esg.py (并发版)
"""
import argparse
import json
import os
import ssl
import time
import urllib.request
from pathlib import Path

import openpyxl

ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Referer': 'https://finance.sina.com.cn/esg/grade.shtml',
}

_PROJECT_ROOT = Path(__file__).parent.parent
DEFAULT_EXCEL_PATH = str(_PROJECT_ROOT / "data" / "msci_input.xlsx")
DEFAULT_OUTPUT_PATH = str(_PROJECT_ROOT / "data" / "msci_output.xlsx")
DEFAULT_PERSON = "your_name"


def get_msci(symbol):
    url = f'https://global.finance.sina.com.cn/api/openapi.php/EsgService.getEsgStockInfo?symbol={symbol}'
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=15, context=ctx) as response:
            data = json.loads(response.read().decode('utf-8'))
            for info in data.get('result', {}).get('data', {}).get('info', []):
                if info.get('agency_name') == 'MSCI':
                    return info.get('esg_score'), info.get('esg_dt')
    except Exception:
        pass
    return None, None


def main():
    parser = argparse.ArgumentParser(description="批量获取 MSCI ESG 评级 (DEPRECATED)")
    parser.add_argument(
        '--person',
        default=os.environ.get('PERSON_NAME', DEFAULT_PERSON),
        help='Excel 中负责人的列值 (环境变量: PERSON_NAME)',
    )
    parser.add_argument('--excel', default=os.environ.get('EXCEL_PATH', DEFAULT_EXCEL_PATH))
    parser.add_argument('--output', default=DEFAULT_OUTPUT_PATH)
    parser.add_argument(
        '--sheet',
        default='2025年度上市公司名单',
        help='Excel sheet 名称',
    )
    args = parser.parse_args()

    print("⚠️  DEPRECATED — 请改用 fetch_msci_esg.py")
    print()

    person = args.person
    cache_key = person.replace(' ', '_')
    cache_file = str(_PROJECT_ROOT / "data" / f"msci_esg_ratings_{cache_key}.json")

    print("1. 读取 Excel...")
    wb = openpyxl.load_workbook(args.excel)
    ws = wb[args.sheet]

    target_stocks = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) > 6 and row[6] == person:
            ts_code = row[1]
            hk_code = row[2]
            if ts_code and ts_code.endswith('.SH'):
                symbol = 'sh' + ts_code.replace('.SH', '')
            elif ts_code and ts_code.endswith('.SZ'):
                symbol = 'sz' + ts_code.replace('.SZ', '')
            else:
                symbol = None
            target_stocks.append({
                'row_idx': row_idx,
                'name': row[0],
                'ts_code': ts_code,
                'hk_code': hk_code,
                'short_name': row[3],
                'symbol': symbol,
            })

    print(f"   [{person}] 股票总数: {len(target_stocks)}")
    stocks_with_sym = [s for s in target_stocks if s['symbol']]
    print(f"   有 API 符号: {len(stocks_with_sym)}")
    no_sym = [s for s in target_stocks if not s['symbol']]
    if no_sym:
        print(f"   无符号 (北交所等): {len(no_sym)}")

    cache = {}
    if os.path.exists(cache_file):
        with open(cache_file, encoding='utf-8') as f:
            cache = json.load(f)
        print(f"   缓存已有: {len(cache)} 条")

    to_fetch = [s for s in stocks_with_sym if s['symbol'].upper() not in cache]
    print(f"   需新获取: {len(to_fetch)} 只")

    if to_fetch:
        print("\n2. 开始获取...")
        fetched = 0
        start = time.time()
        for i, s in enumerate(to_fetch):
            sym = s['symbol']
            rating, date = get_msci(sym)
            cache[sym.upper()] = {'msci_rating': rating, 'msci_date': date}
            if rating:
                fetched += 1
            if (i + 1) % 50 == 0:
                elapsed = time.time() - start
                rate = elapsed / (i + 1)
                eta = rate * (len(to_fetch) - i - 1) if (i + 1) > 0 else 0
                print(f"   {i+1}/{len(to_fetch)} | 成功: {fetched} | ETA: {eta/60:.1f}min")
            time.sleep(0.3)

        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
        print(f"   缓存已保存, 新增成功: {fetched}")

    print("\n3. 回填 Excel...")
    filled = 0
    no_rating = 0
    no_sym_count = 0

    for s in target_stocks:
        row_idx = s['row_idx']
        if not s['symbol']:
            no_sym_count += 1
            continue
        sym_upper = s['symbol'].upper()
        if sym_upper in cache:
            result = cache[sym_upper]
            rating = result.get('msci_rating')
            date = result.get('msci_date')
            if rating and rating != '-':
                ws.cell(row=row_idx, column=5, value=rating)
                ws.cell(row=row_idx, column=6, value=date)
                filled += 1
            else:
                no_rating += 1
        else:
            no_rating += 1

    wb.save(args.output)
    print(f"   保存: {args.output}")
    print(f"   成功填充: {filled}/{len(target_stocks)}")
    print(f"   无 MSCI 评级: {no_rating}")
    print(f"   无 API 符号: {no_sym_count}")

    dist = {}
    for s in target_stocks:
        if s['symbol']:
            r = cache.get(s['symbol'].upper(), {}).get('msci_rating')
            if r and r != '-':
                dist[r] = dist.get(r, 0) + 1

    print("\n=== 评级分布 ===")
    for r in sorted(dist.keys()):
        print(f"  {r}: {dist[r]}")
    print("Done!")


if __name__ == '__main__':
    main()

#!/usr/bin/env python3
"""
从新浪财经ESG API批量获取许哲逸负责的1333只股票的MSCI评级

用法：
    python scripts/fetch_msci_xu.py --excel /path/to/input.xlsx --output /path/to/output.xlsx

环境变量：
    EXCEL_PATH   — 输入Excel文件路径
    OUTPUT_PATH  — 输出Excel文件路径
    CACHE_DIR    — 缓存目录（默认: data/）
"""
import concurrent.futures
import json
import os
import ssl
import time
import urllib.request
from pathlib import Path

import openpyxl

# ── 路径配置（支持环境变量或命令行参数）─────────────────────────────
_PROJECT_ROOT = Path(__file__).parent.parent
_CACHE_DIR = Path(os.environ.get("CACHE_DIR", str(_PROJECT_ROOT / "data")))
_EXCEL_PATH = os.environ.get("EXCEL_PATH", str(Path.home() / "Desktop" / "2026MSCI级别6.1.xlsx"))
_OUTPUT_PATH = os.environ.get("OUTPUT_PATH", str(Path.home() / "Desktop" / "2026MSCI级别6.1_已填充.xlsx"))
CACHE_PATH = str(_CACHE_DIR / "msci_esg_ratings_xu.json")

def _get_paths(args_namespace) -> tuple[str, str]:
    """从参数或环境变量获取路径，文件不存在时给出清晰错误。"""
    excel = args_namespace.excel if args_namespace.excel else _EXCEL_PATH
    output = args_namespace.output if args_namespace.output else _OUTPUT_PATH
    if not os.path.exists(excel):
        raise FileNotFoundError(
            f"输入Excel不存在: {excel}\n"
            f"请通过 --excel 参数指定文件路径，或设置 EXCEL_PATH 环境变量。"
        )
    return excel, output

# === SSL上下文 ===
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Referer': 'https://finance.sina.com.cn/esg/grade.shtml',
}

def normalize_symbol(ts_code: str, hk_code: str) -> str:
    """将tushare格式转换为新浪API格式"""
    if ts_code is None or ts_code == '':
        if hk_code:
            return hk_code.upper().replace('.HK', 'hk').replace('HK', 'hk')
        return None

    ts = ts_code.upper()
    if ts.endswith('.SH'):
        return 'sh' + ts.replace('.SH', '')
    elif ts.endswith('.SZ'):
        return 'sz' + ts.replace('.SZ', '')
    elif ts.endswith('.BJ'):
        # 北交所暂不支持，跳过
        return None
    return None

def get_esg_info(symbol: str, max_retries: int = 2) -> dict:
    """获取单只股票的ESG数据，返回MSCI评级和日期"""
    if not symbol:
        return {'symbol': symbol, 'msci_rating': None, 'msci_date': None}

    url = f'https://global.finance.sina.com.cn/api/openapi.php/EsgService.getEsgStockInfo?symbol={symbol}'
    for attempt in range(max_retries):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=15, context=ctx) as response:
                data = json.loads(response.read().decode('utf-8'))
                info_list = data.get('result', {}).get('data', {}).get('info', [])
                for info in info_list:
                    if info.get('agency_name') == 'MSCI':
                        return {
                            'symbol': symbol.upper(),
                            'msci_rating': info.get('esg_score'),
                            'msci_date': info.get('esg_dt'),
                        }
                return {'symbol': symbol.upper(), 'msci_rating': None, 'msci_date': None}
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(1)
            return {'symbol': symbol.upper(), 'msci_rating': None, 'msci_date': None, 'error': str(e)}
    return {'symbol': symbol.upper(), 'msci_rating': None, 'msci_date': None}

def main():
    # 1. 读取Excel
    print("1. 读取Excel文件...")
    wb = openpyxl.load_workbook(_EXCEL_PATH)
    ws = wb['2025年度上市公司名单']

    xu_stocks = []
    xu_row_indices = []  # Excel行号 (1-indexed)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[6] == '许哲逸':
            ts_code = row[1]
            hk_code = row[2]
            symbol = normalize_symbol(ts_code, hk_code)
            xu_stocks.append({
                'row_idx': row_idx,
                'name': row[0],
                'ts_code': ts_code,
                'hk_code': hk_code,
                'short_name': row[3],
                'symbol': symbol,
            })
            xu_row_indices.append(row_idx)

    print(f"   许哲逸负责股票数: {len(xu_stocks)}")
    stocks_with_symbol = [s for s in xu_stocks if s['symbol']]
    print(f"   有API符号的股票: {len(stocks_with_symbol)}")
    stocks_no_symbol = [s for s in xu_stocks if not s['symbol']]
    if stocks_no_symbol:
        print(f"   无法获取的股票 (无符号): {len(stocks_no_symbol)}")

    # 2. 加载已有缓存数据
    cache = {}
    cache_file = CACHE_PATH
    if os.path.exists(cache_file):
        with open(cache_file, encoding='utf-8') as f:
            cache = json.load(f)
        print(f"   加载缓存: {len(cache)} 条")

    # 3. 确定需要获取的股票
    to_fetch = []
    for s in stocks_with_symbol:
        sym_upper = s['symbol'].upper()
        if sym_upper not in cache:
            to_fetch.append(s)

    print(f"   需要新获取: {len(to_fetch)} 只")
    print(f"   已命中缓存: {len(stocks_with_symbol) - len(to_fetch)} 只")

    if to_fetch:
        print("\n2. 开始获取MSCI数据 (并发5个worker)...")
        fetched = 0
        failed = 0
        start_time = time.time()

        def fetch_one(stock):
            sym_upper = stock['symbol'].upper()
            result = get_esg_info(stock['symbol'])
            return stock, result

        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            futures = {executor.submit(fetch_one, s): s for s in to_fetch}
            done = 0
            for future in concurrent.futures.as_completed(futures):
                stock, result = future.result()
                sym_upper = stock['symbol'].upper()
                cache[sym_upper] = result
                done += 1

                if result.get('msci_rating'):
                    fetched += 1
                else:
                    failed += 1

                # 每50个打印进度
                if done % 50 == 0:
                    elapsed = time.time() - start_time
                    rate = done / elapsed
                    remaining = len(to_fetch) - done
                    eta = remaining / rate if rate > 0 else 0
                    print(f"   进度: {done}/{len(to_fetch)} ({100*done/len(to_fetch):.1f}%) | "
                          f"成功: {fetched} | 失败: {failed} | "
                          f"ETA: {eta/60:.1f}min")

                time.sleep(0.2)  # 避免过快请求

        # 保存缓存
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
        print(f"\n   缓存已保存: {cache_file}")
        print(f"   本轮新增成功: {fetched}")
        print(f"   本轮失败: {failed}")

    # 4. 回填Excel
    print("\n3. 回填Excel数据...")
    filled_count = 0
    no_rating_count = 0
    no_symbol_count = 0

    for s in xu_stocks:
        row_idx = s['row_idx']
        if not s['symbol']:
            no_symbol_count += 1
            continue

        sym_upper = s['symbol'].upper()
        if sym_upper in cache:
            result = cache[sym_upper]
            rating = result.get('msci_rating')
            date = result.get('msci_date')
            if rating and rating != '-':
                ws.cell(row=row_idx, column=5, value=rating)  # MSCI评级
                ws.cell(row=row_idx, column=6, value=date)   # 更新日期
                filled_count += 1
            else:
                no_rating_count += 1
        else:
            no_rating_count += 1

    print(f"   成功填充MSCI评级: {filled_count}")
    print(f"   无MSCI评级: {no_rating_count}")
    print(f"   无API符号(北交所等): {no_symbol_count}")

    # 5. 保存
    print("\n4. 保存文件...")
    wb.save(_OUTPUT_PATH)
    print(f"   已保存: {_OUTPUT_PATH}")

    # 6. 统计总览
    print("\n=== 汇总 ===")
    total = len(xu_stocks)
    print(f"许哲逸负责总数: {total}")
    print(f"成功获取评级: {filled_count} ({100*filled_count/total:.1f}%)")
    print(f"无MSCI评级: {no_rating_count} ({100*no_rating_count/total:.1f}%)")
    print(f"无API符号: {no_symbol_count} ({100*no_symbol_count/total:.1f}%)")

    # 评级分布
    rating_dist = {}
    for s in xu_stocks:
        if s['symbol']:
            sym_upper = s['symbol'].upper()
            if sym_upper in cache:
                r = cache[sym_upper].get('msci_rating', '-')
                if r and r != '-':
                    rating_dist[r] = rating_dist.get(r, 0) + 1
    print("\nMSCI评级分布:")
    for r in sorted(rating_dist.keys()):
        print(f"  {r}: {rating_dist[r]}")

if __name__ == '__main__':
    main()

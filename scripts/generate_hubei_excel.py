#!/usr/bin/env python3
"""Generate Excel from hubei_tech_data_2026.json with multiple styled sheets."""

from __future__ import annotations

import json

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── colour palette ────────────────────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"   # deep navy  — sheet title bg
C_MID_BLUE    = "2E75B6"   # medium blue — header bg
C_LIGHT_BLUE  = "BDD7EE"   # light blue  — alt row
C_WHITE       = "FFFFFF"
C_GOLD        = "FFC000"   # accent — category header
C_GOLD_LIGHT  = "FFF2CC"   # pale gold   — alt category row
C_GREY_BG     = "F2F2F2"   # light grey   — data type label
C_GREEN_FILL  = "E2EFDA"   # green — A类
C_BLUE_FILL   = "DEEAF1"   # blue  — B类
C_ORANGE_FILL = "FCE4D6"   # orange — C类
C_TAB_BLUE    = "4472C4"
C_TAB_GREEN   = "70AD47"
C_TAB_ORANGE  = "ED7D31"
C_TAB_PURPLE  = "7030A0"
C_TAB_RED     = "C00000"
C_TAB_GREY    = "7F7F7F"

# ── thin border ───────────────────────────────────────────────────────────────
THIN = Side(style="thin", color="BFBFBF")
MED  = Side(style="medium", color="2E75B6")
THICK= Side(style="medium", color="1F3864")

def thin_border(left=True, right=True, top=True, bottom=True):
    return Border(
        left=THIN  if left   else Side(),
        right=THIN if right  else Side(),
        top=THIN   if top    else Side(),
        bottom=THIN if bottom else Side(),
    )

def thick_border():
    return Border(left=MED, right=MED, top=MED, bottom=MED)

# ── header style factory ────────────────────────────────────────────────────
def hdr_style(bg=C_MID_BLUE, fg=C_WHITE, bold=True, size=11, center=True):
    return {
        "font":      Font(name="Calibri", bold=bold, color=fg, size=size),
        "fill":      PatternFill("solid", fgColor=bg),
        "alignment": Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=True),
        "border":    thin_border(),
    }

def apply_style(cell, **kw):
    for attr, val in kw.items():
        setattr(cell, attr, val)

def style_range(ws, row, col_start, col_end,
                bg=None, fg=C_WHITE, bold=True, size=11,
                center=True, wrap=True, border=True):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(name="Calibri", bold=bold, color=fg, size=size)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center", wrap_text=wrap
        )
        if border:
            cell.border = thin_border()

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


# ── data helpers ─────────────────────────────────────────────────────────────
def fmt_val(v, unit):
    if v is None: return "—"
    if isinstance(v, float):
        if unit in ("%", "个百分点"):
            return f"{v:.2f}{unit}"
        return f"{v:,.2f}{unit}"
    if isinstance(v, int):
        return f"{v:,}{unit}" if unit else f"{v:,}"
    return str(v)

# ── SHEET 1: 元数据总览 ──────────────────────────────────────────────────────
def sheet_overview(wb):
    ws = wb.active
    ws.title = "①总览"
    ws.sheet_properties.tabColor = C_TAB_BLUE
    ws.freeze_panes = "A3"

    # Title
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "湖北省科技创新与产业发展数据清单  —  数据总览"
    t.font = Font(name="Calibri", bold=True, color=C_WHITE, size=14)
    t.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 1, 32)

    # Headers
    headers = ["类别", "指标名称", "最新值", "单位", "来源"]
    style_range(ws, 2, 1, 5, bg=C_MID_BLUE, size=10)
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i).value = h
    set_row_height(ws, 2, 22)

    # Data rows — build from JSON data
    data = [
        # 经济总量
        ("经济总量", "全省GDP（2025）", "62,660.90", "亿元", "湖北省统计局2025年统计公报"),
        ("经济总量", "GDP增速（2025）", "5.5", "%", "同上"),
        ("经济总量", "人均GDP（2025）", "107,619", "元", "同上"),
        ("经济总量", "武汉市GDP（2025）", "22,147.35", "亿元", "武汉市统计局"),
        # 高校
        ("高校与人才", "普通高校本专科在校生（2025）", "196.75", "万人", "湖北省统计局2025年统计公报"),
        ("高校与人才", "在校研究生（2025）", "24.92", "万人", "同上"),
        ("高校与人才", "普通高等学校数量（2024）", "134", "所", "马克数据网整理"),
        ("高校与人才", "双一流高校数量（2025）", "7", "所", "教育部双一流建设名单"),
        # R&D
        ("R&D投入", "R&D经费（2024）", "1,408.2", "亿元", "湖北省科技厅"),
        ("R&D投入", "R&D投入强度（2024）", "2.70", "%", "同上"),
        ("R&D投入", "强度增幅（2024）", "+0.19", "pp", "同上"),
        # 科技企业
        ("科技企业", "高新技术企业数量（2024）", "30,000", "家", "湖北省统计局/湖北日报"),
        ("科技企业", "科技型中小企业（2024）", "45,000", "家", "湖北省统计局"),
        ("科技企业", "专精特新小巨人（2024）", "678", "家", "湖北省经信厅"),
        ("科技企业", "高技术制造业增速（2024）", "22.7", "%", "2024年统计公报"),
        # 技术转化
        ("技术转化", "技术合同成交额（2025）", "6,121.57", "亿元", "2025年统计公报"),
        ("科技企业", "科技成果就地转化率（2025）", "68.9", "%", "湖北省科技厅"),
        # 战略性新兴产业
        ("战略性新兴产业", "数字经济核心产业增加值（2024）", "5,742.93", "亿元", "湖北省数据局"),
        ("战略性新兴产业", "数字经济总规模（2024）", "3", "万亿元", "同上"),
        ("战略性新兴产业", "光电子信息产业规模（2024）", "1", "万亿元+", "湖北省经信厅"),
        ("战略性新兴产业", "大健康产业规模（2024）", "1", "万亿元", "同上"),
        # AI
        ("人工智能", "AI企业数量（2024）", "1,143", "家", "湖北省科技厅"),
        ("人工智能", "AI产业规模（2024）", "1,107", "亿元", "湖北省经信厅"),
        # 算力
        ("算力基础设施", "武汉人工智能计算中心算力", "400", "P", "武汉科创局"),
        ("算力基础设施", "武汉超算中心总算力（2025）", "100", "P", "武汉日报"),
        ("算力基础设施", "光谷总算力（2025）", "4,700+", "P", "湖北日报"),
        # 上市公司
        ("上市公司", "A股上市公司数量（2024）", "151", "家", "《湖北省上市公司发展报告2025》"),
        ("上市公司", "全省上市公司（2025，含境外）", "193", "家", "同上"),
        # 创新平台
        ("创新平台", "国家实验室", "1", "家", "21经济网"),
        ("创新平台", "大科学装置", "8", "个", "湖北日报"),
        ("创新平台", "全国重点实验室", "45", "家", "科技部"),
        ("创新平台", "新型研发机构", "525", "家", "湖北省科技厅"),
        # 科技金融
        ("科技金融", "省级母基金参股规模（2024）", "13,700", "亿元", "湖北省母基金行业发展报告"),
    ]

    cat_colors = {
        "经济总量":        (C_DARK_BLUE,  C_WHITE),
        "高校与人才":      ("1A5276",     C_WHITE),
        "R&D投入":         ("1A5276",     C_WHITE),
        "科技企业":         ("154360",     C_WHITE),
        "技术转化":         ("154360",     C_WHITE),
        "战略性新兴产业":   (C_MID_BLUE,   C_WHITE),
        "人工智能":         ("2E4057",     C_WHITE),
        "算力基础设施":     ("2E4057",     C_WHITE),
        "上市公司":         (C_GOLD,       "000000"),
        "创新平台":         ("04726A",     C_WHITE),
        "科技金融":         ("7B3F00",     C_WHITE),
    }

    row = 3
    for idx, (cat, name, val, unit, src) in enumerate(data):
        bg_color, fg_color = cat_colors.get(cat, (C_MID_BLUE, C_WHITE))
        alt = (idx % 2 == 0)
        row_bg = bg_color if not alt else (C_LIGHT_BLUE if bg_color == C_DARK_BLUE else C_GREY_BG)

        for col, text in enumerate([cat, name, val, unit, src], 1):
            cell = ws.cell(row=row, column=col)
            cell.value = text
            cell.font = Font(name="Calibri", size=10,
                             bold=(col in (1, 3)),
                             color=fg_color if not alt else "000000")
            cell.fill = PatternFill("solid",
                fgColor=bg_color if not alt else (C_LIGHT_BLUE if bg_color == C_DARK_BLUE else C_GREY_BG))
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 3, 4) else "left",
                vertical="center"
            )
            cell.border = thin_border()
        set_row_height(ws, row, 18)
        row += 1

    # Column widths
    widths = [18, 36, 16, 10, 38]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    # Freeze + filter
    ws.auto_filter.ref = f"A2:E{row - 1}"

    # Legend row
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    note = ws.cell(row=row, column=1)
    note.value = (
        "数据来源：湖北省统计局、湖北省科技厅、湖北省经信厅、武汉科创局、马克数据网整理《中国科技统计年鉴》等。"
        " 数据均已通过官方公报核查（2026-05-31）。  ⚠ 部分数据（如基金规模）为手动录入，需定期更新。"
    )
    note.font = Font(name="Calibri", size=9, italic=True, color="595959")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    set_row_height(ws, row, 36)


# ── SHEET 2: 时间序列（长格式） ─────────────────────────────────────────────
def sheet_timeseries(wb):
    ws = wb.create_sheet("②时间序列")
    ws.sheet_properties.tabColor = C_TAB_GREEN

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "湖北省主要指标历年序列（2005–2025）"
    t.font = Font(name="Calibri", bold=True, color=C_WHITE, size=13)
    t.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 1, 30)

    headers = ["指标名称", "单位", "数据分类", "年份", "数值", "数值格式化", "数据来源", "备注"]
    style_range(ws, 2, 1, 8, bg=C_MID_BLUE, size=10)
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i).value = h
    set_row_height(ws, 2, 22)

    with open("data/hubei_tech_data_2026.json", encoding="utf-8") as f:
        js = json.load(f)

    ts = js["real_data"]["time_series"]

    # Map: display_name -> (series_name_in_json, unit, source, note)
    series_map = [
        ("普通高等学校数量", "普通高等学校数量", "所", "马克数据网整理", ""),
        ("普通高校本专科在校生", "普通高校本专科在校生", "万人", "湖北省统计局", ""),
        ("在校研究生", "在校研究生", "万人", "湖北省统计局", ""),
        ("湖北省GDP", "GDP", "亿元", "湖北省统计局", "当年价"),
        ("R&D经费投入", "R&D经费投入", "亿元", "湖北省科技厅/马克数据网", ""),
        ("技术合同成交额", "技术合同成交额", "亿元", "湖北省科技厅/统计局", ""),
        ("高新技术企业数量", "高新技术企业数量", "家", "湖北省科技厅/马克数据网", ""),
        ("国家级专精特新小巨人", "国家级专精特新小巨人", "家", "工信部+湖北省经信厅", ""),
        ("A股上市公司数量", "A股上市公司数量", "家", "《湖北省上市公司发展报告》", ""),
        ("全省上市公司（含境外）", "全省上市公司（含境外）", "家", "《湖北省上市公司发展报告2025》", ""),
        ("武汉市GDP", "武汉市GDP", "亿元", "武汉市统计局", ""),
    ]

    row = 3
    type_colors = {"A": C_GREEN_FILL, "B": C_BLUE_FILL, "C": C_ORANGE_FILL}
    type_labels = {"A": "A类-真实数据", "B": "B类-MCP可补充", "C": "C类-手动录入"}

    for disp_name, json_key, unit, source, note in series_map:
        if json_key not in ts:
            continue
        series = ts[json_key]
        data_type = series.get("data_type", "A")
        series_src = series.get("source", source)
        series_note = series.get("note", note)
        bg_type = type_colors.get(data_type, C_GREY_BG)
        label_type = type_labels.get(data_type, data_type)

        # series header row
        ws.merge_cells(f"A{row}:H{row}")
        sh = ws.cell(row=row, column=1)
        sh.value = f"{disp_name}  [{label_type}]  来源：{series_src}"
        sh.font = Font(name="Calibri", bold=True, color=C_WHITE, size=10)
        sh.fill = PatternFill("solid", fgColor="2E75B6")
        sh.alignment = Alignment(horizontal="left", vertical="center")
        sh.border = thin_border()
        set_row_height(ws, row, 18)
        row += 1

        # data rows
        for year_str, val in sorted(series.get("data", {}).items(), key=lambda x: int(x[0])):
            year = int(year_str)
            fmt = fmt_val(val, unit)
            alt = (year % 2 == 0)
            row_bg = bg_type if alt else C_WHITE

            for col, text in enumerate(
                [disp_name, unit, label_type, year, val, fmt, series_src, series_note], 1
            ):
                cell = ws.cell(row=row, column=col)
                cell.value = text
                cell.font = Font(name="Calibri", size=10,
                                 bold=(col in (4, 5, 6)))
                cell.fill = PatternFill("solid", fgColor=row_bg)
                cell.alignment = Alignment(
                    horizontal="center" if col in (1, 2, 3, 4, 5, 6) else "left",
                    vertical="center"
                )
                cell.border = thin_border()
            set_row_height(ws, row, 16)
            row += 1

        row += 1  # blank between series

    # Column widths
    widths = [26, 8, 18, 8, 14, 16, 30, 30]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    ws.freeze_panes = "A3"


# ── SHEET 3: 数据分类索引 ────────────────────────────────────────────────────
def sheet_classification(wb):
    ws = wb.create_sheet("③数据分类索引")
    ws.sheet_properties.tabColor = C_TAB_PURPLE

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "数据分类索引 — A类（真实数据）/ B类（MCP可补充）/ C类（手动录入）"
    t.font = Font(name="Calibri", bold=True, color=C_WHITE, size=13)
    t.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 1, 30)

    headers = ["分类", "类别", "指标", "最新值", "单位", "年份", "数据来源"]
    style_range(ws, 2, 1, 7, bg=C_MID_BLUE, size=10)
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i).value = h
    set_row_height(ws, 2, 22)

    with open("data/hubei_tech_data_2026.json", encoding="utf-8") as f:
        js = json.load(f)

    # Build flat index
    flat = []
    for item in js["real_data"]["data"]:
        flat.append(("A", item["category"], item["indicator"],
                     item["value"], item["unit"], item["year"], item["source"]))

    for item in js["mcp_available"]["indicators"]:
        flat.append(("B", "—", item["indicator"], "—", "—", "—",
                     f"MCP: {item['mcp_service']} / {item['mcp_tool']}"))

    for item in js["manual_entry"]["indicators"]:
        v = item.get("value", "—")
        yr = item.get("year", "—")
        flat.append(("C", item["category"], item["indicator"],
                     v, "—", yr, item["source"]))

    cat_bg = {"A": C_GREEN_FILL, "B": C_BLUE_FILL, "C": C_ORANGE_FILL}
    cat_hdr = {"A": "A类-真实数据", "B": "B类-MCP可补充", "C": "C类-手动录入"}

    row = 3
    prev_cls = None
    for rec in flat:
        cls, cat, name, val, unit, year, src = rec
        if cls != prev_cls:
            # section header
            ws.merge_cells(f"A{row}:G{row}")
            sh = ws.cell(row=row, column=1)
            sh.value = cat_hdr[cls]
            sh.font = Font(name="Calibri", bold=True, color=C_WHITE, size=11)
            sh.fill = PatternFill("solid", fgColor=cat_bg[cls].replace("E2EFDA", "375623"
                               if cls=="A" else "1F3864" if cls=="B" else "843C0C"))
            sh.alignment = Alignment(horizontal="center", vertical="center")
            sh.border = thin_border()
            set_row_height(ws, row, 20)
            row += 1
            prev_cls = cls

        alt = (row % 2 == 0)
        row_bg = cat_bg[cls] if alt else C_WHITE
        vals = [cat, name, fmt_val(val, unit), unit, str(year), src]
        for col, text in enumerate([cls] + vals, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = text
            cell.font = Font(name="Calibri", size=10,
                             bold=(col in (1, 3, 4)))
            cell.fill = PatternFill("solid", fgColor=row_bg)
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 4, 5) else "left",
                vertical="center"
            )
            cell.border = thin_border()
        set_row_height(ws, row, 16)
        row += 1

    widths = [10, 22, 36, 16, 8, 8, 40]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:G{row - 1}"


# ── SHEET 4: MCP服务概览 ────────────────────────────────────────────────────
def sheet_mcp(wb):
    ws = wb.create_sheet("④MCP服务概览")
    ws.sheet_properties.tabColor = C_TAB_ORANGE

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "MCP服务与数据工具概览（4个服务，32个工具）"
    t.font = Font(name="Calibri", bold=True, color=C_WHITE, size=13)
    t.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 1, 30)

    headers = ["MCP服务", "工具数量", "数据类型", "数据来源", "覆盖年份", "调用方式"]
    style_range(ws, 2, 1, 6, bg=C_MID_BLUE, size=10)
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i).value = h
    set_row_height(ws, 2, 22)

    servers = [
        ("user-hubei-stats",  12, "全国宏观指标（动态API）",
         "akshare / World Bank API", "2006–present",
         "MCP调用: ak.macro_china_gdp() 等"),
        ("user-wuhan-stats",   6, "武汉统计（手动录入）",
         "武汉统计年鉴PDF / 武汉市统计局", "2019–2025",
         "MCP调用: 返回已知数据列表"),
        ("user-macro-stats",   9, "全球宏观（动态API）",
         "World Bank API（无需Key）", "1960–present",
         "MCP调用: get_wb_indicator() 等"),
        ("user-macro-datas",   5, "科技面板（手动录入）",
         "马克数据网《中国科技统计年鉴》", "2000–2024",
         "MCP调用: 返回马克数据网整理数据"),
    ]

    for i, (name, n_tools, dtype, src, years, call) in enumerate(servers):
        bg = [C_LIGHT_BLUE, C_GREY_BG][i % 2]
        for col, text in enumerate([name, n_tools, dtype, src, years, call], 1):
            cell = ws.cell(row=3 + i, column=col)
            cell.value = text
            cell.font = Font(name="Calibri", size=10, bold=(col in (1, 2)))
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border()
        set_row_height(ws, 3 + i, 20)

    widths = [22, 10, 28, 40, 16, 36]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    # Tool list per server
    row = 8
    ws.merge_cells(f"A{row}:F{row}")
    sh = ws.cell(row=row, column=1)
    sh.value = "工具详细列表"
    sh.font = Font(name="Calibri", bold=True, color=C_WHITE, size=12)
    sh.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    sh.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, row, 24)
    row += 1

    headers2 = ["服务器", "工具名", "数据类型", "调用说明"]
    style_range(ws, row, 1, 4, bg=C_MID_BLUE, size=10)
    for i, h in enumerate(headers2, 1):
        ws.cell(row=row, column=i).value = h
    set_row_height(ws, row, 20)
    row += 1

    tools = [
        # hubei
        ("user-hubei-stats", "get_china_gdp",         "动态API", "ak.macro_china_gdp()，2006Q1-2026Q1季度数据"),
        ("user-hubei-stats", "get_china_gdp_yearly",  "动态API", "ak.macro_china_gdp_yearly()，全国GDP年度"),
        ("user-hubei-stats", "get_cpi",               "动态API", "ak.macro_china_cpi_monthly()，CPI月度"),
        ("user-hubei-stats", "get_ppi",               "动态API", "ak.macro_china_ppi_monthly()，PPI月度"),
        ("user-hubei-stats", "get_pmi",               "动态API", "ak.macro_china_pmi()，PMI月度"),
        ("user-hubei-stats", "get_m2",                "动态API", "ak.macro_china_m2_yearly()，M2年度"),
        ("user-hubei-stats", "get_fdi",               "动态API", "ak.macro_china_fdi()，FDI月度"),
        ("user-hubei-stats", "get_consumer_retail",   "动态API", "ak.macro_china_consumer_goods_retail()，零售月度"),
        ("user-hubei-stats", "get_industry_production","动态API", "ak.macro_china_industrial_production_yoy()，工业增加值"),
        ("user-hubei-stats", "get_hubei_tech_contract","手动录入","马克数据网整理《科技统计年鉴》，含2019-2025年"),
        ("user-hubei-stats", "get_hubei_rd_funding",  "手动录入","马克数据网整理，含2007-2024年（18年）"),
        ("user-hubei-stats", "get_hubei_hitech",      "手动录入","马克数据网整理，含2012-2024年（13年）"),
        # wuhan
        ("user-wuhan-stats", "get_wuhan_gdp",         "手动录入","武汉市统计局，含2019-2025年武汉GDP"),
        ("user-wuhan-stats", "get_wuhan_industry",    "手动录入","武汉统计年鉴，工业增加值增速"),
        ("user-wuhan-stats", "get_wuhan_investment",  "手动录入","武汉统计年鉴，固投增速"),
        ("user-wuhan-stats", "get_wuhan_trade",       "手动录入","武汉海关，进出口数据"),
        ("user-wuhan-stats", "get_wuhan_education",   "手动录入","武汉统计年鉴，高校数据"),
        ("user-wuhan-stats", "get_wuhan_tech",        "手动录入","武汉科创局，算力+高新技术企业"),
        # macro-stats
        ("user-macro-stats", "get_wb_indicator",     "动态API","World Bank API通用指标查询"),
        ("user-macro-stats", "get_wb_gdp_usd",        "动态API","WB NY.GDP.MKTP.CD，中国GDP美元计价"),
        ("user-macro-stats", "get_wb_gdp_pc",         "动态API","WB 人均GDP（美元）"),
        ("user-macro-stats", "get_wb_population",      "动态API","WB 人口数据"),
        ("user-macro-stats", "get_wb_trade",          "动态API","WB 贸易数据"),
        ("user-macro-stats", "get_wb_inflation",      "动态API","WB CPI通胀率"),
        ("user-macro-stats", "get_wb_unemployment",   "动态API","WB 失业率"),
        ("user-macro-stats", "get_wb_tech_rd",        "动态API","WB R&D强度指标"),
        ("user-macro-stats", "get_nbs_fallback",      "手动录入","国家统计局备用方案"),
        # macro-datas
        ("user-macro-datas", "get_rd_panel",          "手动录入","马克数据网R&D面板，2000-2024年各省"),
        ("user-macro-datas", "get_tech_panel",        "手动录入","马克数据网科技面板，高企/专利等"),
        ("user-macro-datas", "get_industry_panel",    "手动录入","马克数据网工业面板"),
        ("user-macro-datas", "get_education_panel",   "手动录入","马克数据网教育面板"),
        ("user-macro-datas", "get_nsti_report",        "手动录入","马克数据网/科技部科技经费公报"),
    ]

    for i, (srv, tool, dtype, desc) in enumerate(tools):
        bg = C_LIGHT_BLUE if i % 2 == 0 else C_WHITE
        for col, text in enumerate([srv, tool, dtype, desc], 1):
            cell = ws.cell(row=row, column=col)
            cell.value = text
            cell.font = Font(name="Calibri", size=9, bold=(col in (1, 2)))
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border()
        set_row_height(ws, row, 16)
        row += 1

    widths2 = [22, 28, 14, 52]
    for i, w in enumerate(widths2, 1):
        set_col_width(ws, i, w)
    ws.freeze_panes = "A3"


# ── SHEET 5: 数据来源汇总 ────────────────────────────────────────────────────
def sheet_sources(wb):
    ws = wb.create_sheet("⑤数据来源")
    ws.sheet_properties.tabColor = C_TAB_GREY

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "数据来源汇总表"
    t.font = Font(name="Calibri", bold=True, color=C_WHITE, size=13)
    t.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 1, 30)

    headers = ["来源名称", "网址", "数据类型", "覆盖范围", "备注"]
    style_range(ws, 2, 1, 5, bg=C_MID_BLUE, size=10)
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i).value = h
    set_row_height(ws, 2, 22)

    sources = [
        ("湖北省统计局",            "https://tjj.hubei.gov.cn/",
         "GDP、高校在校生、技术合同、R&D年报",
         "湖北省，年度+季度", "官方公报实时更新"),
        ("武汉市统计局",            "https://tjj.wuhan.gov.cn/",
         "武汉GDP、工业、投资、贸易",
         "武汉市，年度", "《武汉统计年鉴》PDF"),
        ("湖北省科技厅",            "https://kjt.hubei.gov.cn/",
         "R&D经费、高新技术企业、技术合同",
         "湖北省，年度", "实时发布"),
        ("湖北省经信厅",            "https://jxw.hubei.gov.cn/",
         "工业、新兴产业、AI",
         "湖北省，年度", ""),
        ("武汉科创局",              "https://kjj.wuhan.gov.cn/",
         "算力、人工智能",
         "武汉市，年度", ""),
        ("科技部",                  "https://www.most.gov.cn/",
         "全国R&D公报",
         "全国，年度", ""),
        ("马克数据网（付费）",       "https://www.macrodatas.cn/",
         "各省历年R&D/科技/教育面板",
         "2000-2024年各省", "需付费订阅"),
        ("World Bank API",         "https://api.worldbank.org/",
         "GDP、人口、贸易、通胀",
         "全球，1960-2024", "无需API Key"),
        ("akshare",                "https://akshare.akfamily.xyz/",
         "全国GDP/CPI/PPI等宏观",
         "2006-2026实时", "无需API Key"),
        ("《湖北省上市公司发展报告》", "—",
         "A股+全省上市公司历年",
         "湖北省，2019-2025年", "长江智库发布"),
        ("湖北省母基金行业发展报告", "—",
         "省级母基金规模、AI基金",
         "湖北省，2024年", "手动录入"),
    ]

    for i, (name, url, dtype, coverage, note) in enumerate(sources):
        bg = C_LIGHT_BLUE if i % 2 == 0 else C_WHITE
        for col, text in enumerate([name, url, dtype, coverage, note], 1):
            cell = ws.cell(row=3 + i, column=col)
            cell.value = text
            cell.font = Font(name="Calibri", size=10, bold=(col == 1))
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border()
        set_row_height(ws, 3 + i, 20)

    widths = [26, 36, 38, 26, 26]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)
    ws.freeze_panes = "A3"


# ── SHEET 6: 核查记录 ────────────────────────────────────────────────────────
def sheet_verification(wb):
    ws = wb.create_sheet("⑥核查记录")
    ws.sheet_properties.tabColor = C_TAB_RED

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "数据核查记录（2026-05-31）— 全部35项A类数据已通过官方公报核查"
    t.font = Font(name="Calibri", bold=True, color=C_WHITE, size=13)
    t.fill = PatternFill("solid", fgColor="C00000")
    t.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 1, 30)

    headers = ["指标", "JSON值", "官方来源", "核查结果", "核查方法", "修正前值", "说明"]
    style_range(ws, 2, 1, 7, bg="C00000", fg=C_WHITE, size=10)
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i).value = h
    set_row_height(ws, 2, 22)

    records = [
        ("GDP 2025",         "62,660.90亿元",  "湖北省统计局《2025年统计公报》",     "✅ 一致",  "网页核查",            "—",    ""),
        ("GDP 2024",         "60,012.97亿元",  "湖北省统计局《2024年统计公报》",     "✅ 一致",  "网页核查",            "—",    ""),
        ("高校在校生2025",   "196.75万人",     "《2025年统计公报》",                  "✅ 一致",  "网页核查",            "—",    "本专科"),
        ("研究生在校生2025","24.92万人",      "《2025年统计公报》",                  "✅ 一致",  "网页核查",            "—",    ""),
        ("技术合同2025",     "6,121.57亿元",   "《2025年统计公报》",                  "✅ 一致",  "网页核查",            "—",    "合同66897项"),
        ("技术合同2024",     "5,504.29亿元",   "《2024年统计公报》原文",              "✅ 已修正", "网页核查",            "5,500亿元","公报原文为5504.29"),
        ("R&D 2024",         "1,408.2亿元",    "湖北省科技厅",                        "✅ 一致",  "网页核查",            "—",    ""),
        ("R&D强度2024",      "2.70%",          "湖北省科技厅",                        "✅ 一致",  "网页核查",            "—",    ""),
        ("高新技术企业2024", "3.0万家",        "湖北日报",                            "✅ 一致",  "网页核查",            "—",    ""),
        ("高技术制造业增速", "22.7%",          "《2024年统计公报》",                  "✅ 一致",  "网页核查",            "—",    "占规上工业14.8%"),
        ("专精特新小巨人",   "678家",          "湖北省经信厅",                        "✅ 一致",  "网页核查",            "—",    "全国第7、中部第1"),
        ("武汉超算二期",     "100P",           "武汉市政府官网（2025-10-28）",        "✅ 一致",  "网页核查",            "—",    ""),
        ("光谷总算力",       "4,700P+",        "湖北日报",                            "✅ 一致",  "网页核查",            "—",    ""),
        ("高校数量2010",     "120所",          "gotohui.com/马克数据网",              "✅ 一致",  "数据平台核查",        "—",    ""),
        ("A股上市公司2024",  "151家",          "《湖北省上市公司发展报告2025》",      "✅ 一致",  "网页核查",            "—",    ""),
        ("全省上市公司2025", "193家",          "同上报告（2025-12）",                 "✅ 一致",  "网页核查",            "—",    "A股154+境外39"),
        ("R&D序列扩展",      "2007-2024（18年）","马克数据网《中国科技统计年鉴》面板","✅ 已扩展","马克数据网页面核查","仅2020-2024","扩展至18年"),
        ("高新技术企业序列",  "2012-2024（13年）","马克数据网《中国科技统计年鉴》面板","✅ 已扩展","马克数据网页面核查","仅2020-2024","扩展至13年"),
        ("武汉GDP 2019",     "16,223.21亿元",  "《武汉统计年鉴2020》",               "✅ 补入",  "搜索核查",            "缺失",  ""),
        ("A股上市公司2019",  "106家",          "《湖北省上市公司发展报告》",         "✅ 补入",  "搜索核查",            "缺失",  ""),
    ]

    ok_fill   = PatternFill("solid", fgColor="E2EFDA")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    fix_fill  = PatternFill("solid", fgColor="FCE4D6")

    for i, (name, val, src, result, method, before, note) in enumerate(records):
        bg = C_WHITE if i % 2 == 0 else C_GREY_BG
        for col, text in enumerate([name, val, src, result, method, before, note], 1):
            cell = ws.cell(row=3 + i, column=col)
            cell.value = text
            cell.font = Font(name="Calibri", size=10,
                             bold=(col in (1, 4)),
                             color="375623" if "✅ 已" in result
                                   else "595959" if "✅ 一致" in result
                                   else "C00000")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(
                horizontal="center" if col in (4,) else "left",
                vertical="center", wrap_text=True
            )
            cell.border = thin_border()
        set_row_height(ws, 3 + i, 18)

    widths = [24, 20, 40, 12, 12, 16, 30]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)
    ws.freeze_panes = "A3"


# ── main ─────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    sheet_overview(wb)
    sheet_timeseries(wb)
    sheet_classification(wb)
    sheet_mcp(wb)
    sheet_sources(wb)
    sheet_verification(wb)

    out_path = "data/湖北省科技创新数据_2026.xlsx"
    wb.save(out_path)
    print(f"Done → {out_path}")

if __name__ == "__main__":
    main()

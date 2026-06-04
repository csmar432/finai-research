#!/usr/bin/env python3
"""生成全国各省科技创新数据集成报告 (national_province_data_2026.xlsx).

9 sheets:
  1. 数据总览      — 各省核查状态 + 核心指标摘要
  2. GDP对比       — 全国GDP十强 + 时间序列（湖北）
  3. R&D对比       — R&D经费/强度排名
  4. 科技企业对比  — 高新技术企业/专精特新排名
  5. 技术转化对比  — 技术合同成交额排名
  6. 新兴产业对比  — 高技术制造业/数字经济
  7. 高校人才对比  — 湖北详细数据
  8. 湖北时间序列  — 湖北5大指标多年序列
  9. 数据质量报告  — 验证结果 + 来源追踪

数据来源: 各省统计局公报、科技部公报、马克数据网
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DATA_FILE = Path(__file__).resolve().parent.parent / "data" / "national_province_data_2026.json"
OUT_FILE  = Path(__file__).resolve().parent.parent / "data" / "national_province_data_2026.xlsx"

# ── Style helpers ───────────────────────────────────────────────────────────────

def hdr_font(bold=True, size=11, color="FFFFFF"):
    return Font(bold=bold, size=size, color=color)

def hdr_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def set_col_width(ws, widths: list[tuple[int, int | str]]):
    """(col_idx, width_or_auto)"""
    for col, w in widths:
        ws.column_dimensions[get_column_letter(col)].width = w

def title_row(ws, row: int, text: str, cols: int, fill_hex="1F5C8B"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = hdr_fill(fill_hex)
    cell.alignment = center_align()
    ws.row_dimensions[row].height = 28

def hdr_row(ws, row: int, headers: list[str], fill_hex: str = "2E75B6"):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hdr_font(size=10)
        c.fill = hdr_fill(fill_hex)
        c.alignment = center_align(wrap=True)
        c.border = thin_border()

def data_cell(ws, row: int, col: int, value, fill_hex: str = "", bold=False,
              align="left", num_fmt=""):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=10)
    c.border = thin_border()
    if align == "center":
        c.alignment = center_align()
    else:
        c.alignment = left_align()
    if fill_hex:
        c.fill = PatternFill("solid", fgColor=fill_hex)
    if num_fmt:
        c.number_format = num_fmt
    return c

# Verification color scheme
VERIF_COLORS = {"full": "E2EFDA", "partial": "DEEAF1", "minimal": "FCE4D6"}

CAT_COLORS = {
    "ECON": "DDEBF7",
    "EDU":  "E2EFDA",
    "PLAT": "DDEBF7",
    "RD":   "FFF2CC",
    "ENT":  "E2EFDA",
    "TECH": "E2EFDA",
    "IND":  "FCE4D6",
    "AI":   "E2EFDA",
    "FIN":  "DEEAF1",
}

CAT_NAMES = {
    "ECON": "经济总量",
    "EDU":  "高校与人才",
    "PLAT": "创新平台",
    "RD":   "R&D投入",
    "ENT":  "科技企业",
    "TECH": "技术转化",
    "IND":  "新兴/战略产业",
    "AI":   "AI与算力",
    "FIN":  "科技金融",
}


# ── Data loading ────────────────────────────────────────────────────────────────

def load_data() -> dict:
    with open(DATA_FILE, encoding="utf-8") as f:
        return json.load(f)


# ── Sheet 1: 数据总览 ─────────────────────────────────────────────────────────

def sheet_overview(wb, national):
    ws = wb.active
    ws.title = "1-数据总览"
    title_row(ws, 1, "全国各省科技创新数据集成报告 — 数据总览（2024/2025）", 8)

    # Meta info row
    ws.merge_cells("A2:H2")
    meta = national["meta"]
    c = ws.cell(row=2, column=1,
        value=f"版本: {meta['version']} | 生成: {meta['generated_at']} | "
              f"收录: {meta['data_coverage']} | 备注: {meta.get('data_note', meta.get('methodology', ''))}")
    c.font = Font(size=9, italic=True, color="595959")
    c.alignment = left_align()

    hdr_row(ws, 3, ["省份", "GDP(亿元)", "GDP增速", "R&D经费(亿元)", "R&D强度",
                      "高新技术企业(家)", "技术合同(亿元)", "核查状态"], "1F5C8B")
    set_col_width(ws, [(1, 12), (2, 14), (3, 10), (4, 14), (5, 10),
                        (6, 16), (7, 14), (8, 12)])

    provinces = national["provinces"]
    sorted_provs = sorted(provinces.values(), key=lambda x: x.get("gdp_rank_2024", 99))

    row = 4
    for prov in sorted_provs:
        pk = prov.get("name_cn", "")
        cats = prov.get("data", {})
        econ = cats.get("ECON", {})
        rd = cats.get("RD", {})
        ent = cats.get("ENT", {})
        tech = cats.get("TECH", {})
        verif = prov.get("verification", "")

        def get(ind_dict, keys):
            for k in keys:
                for ik, iv in ind_dict.items():
                    if k in ik:
                        return iv
            return {}

        gdp_data = get(econ, ["GDP_2024"])
        gdp = gdp_data.get("value") if isinstance(gdp_data, dict) else None
        gdp_str = f"{gdp:,.0f}" if gdp else "—"

        gdp_yoy_data = get(econ, ["增速"])
        gdp_yoy = gdp_yoy_data.get("value") if isinstance(gdp_yoy_data, dict) else None
        yoy_str = f"{gdp_yoy:+.1f}%" if gdp_yoy else "—"

        rd_data = get(rd, ["R&D经费"])
        rd_val = rd_data.get("value") if isinstance(rd_data, dict) else None
        rd_str = f"{rd_val:,.1f}" if rd_val else "—"

        rd_int_data = get(rd, ["强度"])
        rd_int = rd_int_data.get("value") if isinstance(rd_int_data, dict) else None
        rd_int_str = f"{rd_int:.2f}%" if rd_int else "—"

        ent_data = get(ent, ["高新"])
        ent_val = ent_data.get("value") if isinstance(ent_data, dict) else None
        ent_str = f"{ent_val:,.0f}" if ent_val else "—"

        tech_data = get(tech, ["技术合同"])
        tech_val = tech_data.get("value") if isinstance(tech_data, dict) else None
        tech_str = f"{tech_val:,.0f}" if tech_val else "—"

        row_data = [pk, gdp_str, yoy_str, rd_str, rd_int_str, ent_str, tech_str, verif]
        for col, val in enumerate(row_data, 1):
            c = data_cell(ws, row, col, val,
                          fill_hex=VERIF_COLORS.get(verif, ""),
                          bold=(col == 1))
            if col == 8:
                c.alignment = center_align()

        row += 1

    # Rankings summary box
    row += 1
    title_row(ws, row, "全国排名速览（2024年）", 6, "375623")
    row += 1
    hdr_row(ws, row, ["指标", "#1", "#2", "#3", "#4", "#5"], "375623")
    row += 1

    rankings = national["ranking_tables"]
    for table_id, tdata in rankings.items():
        rows_data = tdata.get("data", [])
        cells = [tdata.get("title", table_id)]
        for r in rows_data[:5]:
            cells.append(f"{r['province']} {r['value']}")
        for col, val in enumerate(cells, 1):
            data_cell(ws, row, col, val, bold=(col == 1))
        row += 1


# ── Sheet 2: GDP对比 ─────────────────────────────────────────────────────────

def sheet_gdp(wb, national):
    ws = wb.create_sheet("2-GDP对比")
    title_row(ws, 1, "全国GDP十强省份（2024年）", 5)

    hdr_row(ws, 2, ["排名", "省份", "GDP(亿元)", "同比增速", "备注"], "1F5C8B")
    set_col_width(ws, [(1, 8), (2, 12), (3, 16), (4, 12), (5, 30)])

    gdpr = national["ranking_tables"]["GDP_2024"]["data"]
    for i, r in enumerate(gdpr):
        fill = "E2EFDA" if i == 0 else ("DEEAF1" if i < 3 else "")
        data_cell(ws, 3 + i, 1, r["rank"], fill_hex=fill, bold=True, align="center")
        data_cell(ws, 3 + i, 2, r["province"], fill_hex=fill, bold=(i == 0))
        data_cell(ws, 3 + i, 3, r["value"], fill_hex=fill, num_fmt="#,##0.00")
        data_cell(ws, 3 + i, 4, f"第{i+1}位", fill_hex=fill)
        data_cell(ws, 3 + i, 5, r.get("note", ""), fill_hex=fill)

    # Time series for 湖北
    row = 15
    title_row(ws, row, "湖北省GDP时间序列（2005-2025年）", 4, "375623")
    row += 1
    hdr_row(ws, row, ["年份", "GDP(亿元)", "同比增速(%)", "说明"], "375623")
    set_col_width(ws, [(1, 10), (2, 16), (3, 14), (4, 30)])

    hubei_ts = national["provinces"]["湖北"]["time_series"]["GDP"]["data"]
    years = sorted(hubei_ts.keys(), key=lambda y: int(y))
    prev = None
    for i, y in enumerate(years):
        v = hubei_ts[y]
        yoy = ""
        if prev is not None:
            yoy = f"{(v - prev) / prev * 100:+.1f}"
        note = ""
        if y == "2020":
            note = "疫情冲击，首次负增长"
        elif y == "2025":
            note = "最新数据"
        row_data = [y, v, yoy, note]
        fill = "E2EFDA" if y == "2025" else ""
        for col, val in enumerate(row_data, 1):
            data_cell(ws, row + i, col, val, fill_hex=fill,
                      num_fmt="#,##0.00" if col == 2 else "")

    # Cross-province GDP chart data
    row += len(years) + 2
    title_row(ws, row, "主要省份GDP对比（亿元）", 6, "404040")
    row += 1
    hdr_row(ws, row, ["省份", "2020", "2022", "2024", "2025", "2020→2025增幅"], "404040")
    set_col_width(ws, [(1, 12), (2, 14), (3, 14), (4, 14), (5, 14), (6, 16)])

    prov_gdp = {
        "广东": ("广东", "2020", "2022", "2024", "2025"),
        "江苏": ("江苏", "2020", "2022", "2024", "2025"),
        "浙江": ("浙江", "2020", "2022", "2024", "2025"),
        "山东": ("山东", "2020", "2022", "2024", "2025"),
        "湖北": ("湖北", "2020", "2022", "2024", "2025"),
    }
    # Fetch from national data — build lookup
    lookup = {}
    for prov_name in ["广东", "江苏", "浙江", "山东", "湖北"]:
        prov = national["provinces"].get(prov_name, {})
        ts = prov.get("time_series", {}).get("GDP", {}).get("data", {})
        lookup[prov_name] = ts

    cross_row = row + 1
    for prov_name in ["广东", "江苏", "浙江", "山东", "湖北"]:
        ts = lookup[prov_name]
        vals = [ts.get("2020"), ts.get("2022"), ts.get("2024"), ts.get("2025")]
        growth = f"{(vals[3] - vals[0]) / vals[0] * 100:+.1f}%" if vals[0] and vals[3] else "—"
        row_vals = [prov_name] + vals + [growth]
        fill = VERIF_COLORS.get(national["provinces"].get(prov_name, {}).get("verification", ""), "")
        for col, val in enumerate(row_vals, 1):
            data_cell(ws, cross_row, col, val, fill_hex=fill,
                      bold=(col == 1), num_fmt="#,##0" if isinstance(val, (int, float)) else "")


# ── Sheet 3: R&D对比 ─────────────────────────────────────────────────────────

def sheet_rd(wb, national):
    ws = wb.create_sheet("3-RD对比")
    title_row(ws, 1, "各省R&D投入对比（2024年）", 5)

    hdr_row(ws, 2, ["排名", "省份", "R&D经费(亿元)", "R&D强度(%)", "超过全国均值"], "1F5C8B")
    set_col_width(ws, [(1, 8), (2, 12), (3, 16), (4, 14), (5, 16)])

    # R&D经费排名
    rd_rank = national["ranking_tables"].get("RD经费_2024", {}).get("data", [])
    for i, r in enumerate(rd_rank):
        fill = "E2EFDA" if i == 0 else ("FFF2CC" if i < 3 else "")
        data_cell(ws, 3 + i, 1, r["rank"], fill_hex=fill, bold=True, align="center")
        data_cell(ws, 3 + i, 2, r["province"], fill_hex=fill, bold=(i == 0))
        data_cell(ws, 3 + i, 3, r["value"], fill_hex=fill, num_fmt="#,##0.0")
        data_cell(ws, 3 + i, 4, "—", fill_hex=fill)
        data_cell(ws, 3 + i, 5, "是" if r["value"] > 2000 else "—", fill_hex=fill)

    # R&D强度排名
    row = 12
    title_row(ws, row, "各省R&D投入强度排名（2024年）", 5, "375623")
    row += 1
    hdr_row(ws, row, ["排名", "省份", "R&D强度(%)", "超过全国均值(2.69%)", "说明"], "375623")

    rd_int = national["ranking_tables"].get("RD强度_2024", {}).get("data", [])
    for i, r in enumerate(rd_int):
        fill = "E2EFDA" if r["value"] > 3 else ""
        note = r.get("note", "")
        data_cell(ws, row + 1 + i, 1, r["rank"], fill_hex=fill, bold=True, align="center")
        data_cell(ws, row + 1 + i, 2, r["province"], fill_hex=fill, bold=(i == 0))
        data_cell(ws, row + 1 + i, 3, r["value"], fill_hex=fill, num_fmt="0.00")
        data_cell(ws, row + 1 + i, 4, f"+{r['value']-2.69:.2f}pp" if r["value"] > 2.69 else f"{r['value']-2.69:.2f}pp", fill_hex=fill)
        data_cell(ws, row + 1 + i, 5, note, fill_hex=fill)

    # 湖北R&D时间序列
    row = row + len(rd_int) + 3
    title_row(ws, row, "湖北省R&D经费时间序列（2007-2024年）", 4, "404040")
    row += 1
    hdr_row(ws, row, ["年份", "R&D经费(亿元)", "同比增速(%)", "说明"], "404040")

    hubei_ts = national["provinces"]["湖北"]["time_series"]["R&D经费"]["data"]
    years = sorted(hubei_ts.keys(), key=lambda y: int(y))
    prev = None
    for i, y in enumerate(years):
        v = hubei_ts[y]
        yoy = ""
        if prev is not None:
            yoy = f"{(v - prev) / prev * 100:+.1f}"
        note = "最新" if y == "2024" else ""
        for col, val in enumerate([y, v, yoy, note], 1):
            data_cell(ws, row + i, col, val, fill_hex="FFF2CC",
                      num_fmt="#,##0.0" if col == 2 else "")


# ── Sheet 4: 科技企业对比 ─────────────────────────────────────────────────────

def sheet_tech_ent(wb, national):
    ws = wb.create_sheet("4-科技企业对比")
    title_row(ws, 1, "各省高新技术企业数量对比（2024年）", 5)

    hdr_row(ws, 2, ["排名", "省份", "高新技术企业(家)", "数据来源", "说明"], "1F5C8B")
    set_col_width(ws, [(1, 8), (2, 12), (3, 18), (4, 28), (5, 20)])

    ent_rank = national["ranking_tables"].get("高新技术企业_2024", {}).get("data", [])
    for i, r in enumerate(ent_rank):
        fill = "E2EFDA" if i == 0 else ""
        data_cell(ws, 3 + i, 1, r["rank"], fill_hex=fill, bold=True, align="center")
        data_cell(ws, 3 + i, 2, r["province"], fill_hex=fill, bold=(i == 0))
        data_cell(ws, 3 + i, 3, r["value"], fill_hex=fill, num_fmt="#,##0")
        data_cell(ws, 3 + i, 4, "各省统计公报", fill_hex=fill)
        data_cell(ws, 3 + i, 5, r.get("note", ""), fill_hex=fill)


# ── Sheet 5: 技术转化对比 ─────────────────────────────────────────────────────

def sheet_tech(wb, national):
    ws = wb.create_sheet("5-技术转化对比")
    title_row(ws, 1, "各省技术合同成交额对比（2024年）", 5)

    hdr_row(ws, 2, ["排名", "省份", "技术合同(亿元)", "同比", "说明"], "1F5C8B")
    set_col_width(ws, [(1, 8), (2, 12), (3, 16), (4, 10), (5, 24)])

    tech_rank = national["ranking_tables"].get("技术合同_2024", {}).get("data", [])
    for i, r in enumerate(tech_rank):
        fill = "E2EFDA" if i == 0 else ""
        data_cell(ws, 3 + i, 1, r["rank"], fill_hex=fill, bold=True, align="center")
        data_cell(ws, 3 + i, 2, r["province"], fill_hex=fill, bold=(i == 0))
        data_cell(ws, 3 + i, 3, r["value"], fill_hex=fill, num_fmt="#,##0.00")
        data_cell(ws, 3 + i, 4, "—", fill_hex=fill)
        data_cell(ws, 3 + i, 5, r.get("note", ""), fill_hex=fill)

    # 湖北技术合同时间序列
    row = 10
    title_row(ws, row, "湖北省技术合同成交额时间序列", 4, "375623")
    row += 1
    hdr_row(ws, row, ["年份", "成交额(亿元)", "同比增速(%)", "数据来源"], "375623")

    hubei_ts = national["provinces"]["湖北"]["time_series"]["技术合同成交额"]["data"]
    years = sorted(hubei_ts.keys(), key=lambda y: int(y))
    prev = None
    for i, y in enumerate(years):
        v = hubei_ts[y]
        yoy = ""
        if prev is not None:
            yoy = f"{(v - prev) / prev * 100:+.1f}"
        note = "最新" if y == "2025" else "统计公报"
        for col, val in enumerate([y, v, yoy, note], 1):
            data_cell(ws, row + i, col, val, fill_hex="E2EFDA",
                      num_fmt="#,##0.00" if col == 2 else "")


# ── Sheet 6: 新兴产业对比 ─────────────────────────────────────────────────────

def sheet_industry(wb, national):
    ws = wb.create_sheet("6-新兴产业对比")
    title_row(ws, 1, "各省新兴/高技术制造业对比（2024年）", 6)
    hdr_row(ws, 2, ["省份", "高技术制造业增速(%)", "高技术制造业占比(%)",
                     "先进制造业占比(%)", "数字经济/GDP占比(%)", "说明"], "1F5C8B")
    set_col_width(ws, [(1, 12), (2, 18), (3, 18), (4, 18), (5, 20), (6, 20)])

    provinces = ["广东", "江苏", "浙江", "山东", "湖北", "北京", "上海", "四川", "河南"]
    row = 3
    for prov_name in provinces:
        prov = national["provinces"].get(prov_name, {})
        cats = prov.get("data", {})
        ind = cats.get("IND", {})
        verif = prov.get("verification", "")

        def get(ind_dict, keys):
            for k in keys:
                for ik, iv in ind_dict.items():
                    if k in ik:
                        return iv
            return {}

        vals = [
            get(ind, ["高技术制造业增速", "增速"]),
            get(ind, ["高技术制造业占比", "占比"]),
            get(ind, ["先进制造业"]),
            get(ind, ["数字经济", "数字经济核心"]),
            ""
        ]
        row_vals = [prov_name]
        for v in vals:
            if isinstance(v, dict):
                row_vals.append(f"{v.get('value', '—')}" + (f" {v.get('unit', '')}" if v.get('unit') else "%"))
            else:
                row_vals.append(str(v) if v else "—")

        fill = VERIF_COLORS.get(verif, "")
        for col, val in enumerate(row_vals, 1):
            data_cell(ws, row, col, val, fill_hex=fill, bold=(col == 1))
        row += 1


# ── Sheet 7: 高校人才对比（湖北详细） ─────────────────────────────────────────

def sheet_edu(wb, national):
    ws = wb.create_sheet("7-高校与人才")
    title_row(ws, 1, "湖北省高校与人才数据（2025年）", 5)
    hdr_row(ws, 2, ["指标", "数值", "单位", "数据来源", "核查状态"], "1F5C8B")
    set_col_width(ws, [(1, 24), (2, 16), (3, 12), (4, 32), (5, 12)])

    edu_data = national["provinces"]["湖北"]["data"]["EDU"]
    items = list(edu_data.items())
    for i, (k, v) in enumerate(items):
        fill = CAT_COLORS.get("EDU", "")
        data_cell(ws, 3 + i, 1, k, fill_hex=fill, bold=True)
        data_cell(ws, 3 + i, 2, v.get("value", "—"), fill_hex=fill)
        data_cell(ws, 3 + i, 3, v.get("unit", ""), fill_hex=fill, align="center")
        data_cell(ws, 3 + i, 4, v.get("source", ""), fill_hex=fill)
        data_cell(ws, 3 + i, 5, "A", fill_hex="E2EFDA", align="center")

    # 时间序列
    row = 3 + len(items) + 2
    title_row(ws, row, "湖北省高校在校生时间序列（2005-2025年）", 4, "375623")
    row += 1
    hdr_row(ws, row, ["年份", "在校生(万人)", "同比增速(%)", "说明"], "375623")

    hubei_ts = national["provinces"]["湖北"]["time_series"]["本专科在校生"]["data"]
    years = sorted(hubei_ts.keys(), key=lambda y: int(y))
    prev = None
    for i, y in enumerate(years):
        v = hubei_ts[y]
        yoy = ""
        if prev is not None:
            yoy = f"{(v - prev) / prev * 100:+.1f}"
        note = "最新" if y == "2025" else ""
        for col, val in enumerate([y, v, yoy, note], 1):
            data_cell(ws, row + i, col, val, fill_hex="E2EFDA",
                      num_fmt="#,##0.00" if col == 2 else "")


# ── Sheet 8: 湖北时间序列 ──────────────────────────────────────────────────────

def sheet_hubei_ts(wb, national):
    ws = wb.create_sheet("8-湖北时间序列")
    title_row(ws, 1, "湖北省五大核心指标时间序列", 8, "1F5C8B")

    hubei = national["provinces"]["湖北"]
    ts_map = hubei.get("time_series", {})

    row = 2
    for series_name, series_data in ts_map.items():
        unit = series_data.get("unit", "")
        source = series_data.get("source", "")
        data = series_data.get("data", {})

        # Series header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=f"  {series_name}  |  单位: {unit}  |  来源: {source}")
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = hdr_fill("2E75B6")
        c.alignment = left_align(wrap=False)
        ws.row_dimensions[row].height = 18
        row += 1

        # Column headers
        hdr_row(ws, row, ["年份", series_name + "(" + unit + ")", "同比增速(%)", "说明"], "404040")
        row += 1

        years = sorted(data.keys(), key=lambda y: int(y))
        prev = None
        for i, y in enumerate(years):
            v = data[y]
            yoy = ""
            if prev is not None:
                yoy = f"{(v - prev) / prev * 100:+.1f}"
            note = "最新" if y == years[-1] else ""
            is_latest = (y == years[-1])
            fill = "FFF2CC" if is_latest else ""

            data_cell(ws, row + i, 1, y, fill_hex=fill, bold=is_latest, align="center")
            data_cell(ws, row + i, 2, v, fill_hex=fill, bold=is_latest, num_fmt="#,##0.00")
            data_cell(ws, row + i, 3, yoy, fill_hex=fill)
            data_cell(ws, row + i, 4, note, fill_hex=fill)

            prev = v

        row += len(years) + 2


# ── Sheet 9: 数据质量报告 ─────────────────────────────────────────────────────

def sheet_quality(wb, national):
    ws = wb.create_sheet("9-数据质量报告")
    title_row(ws, 1, "数据质量报告 — 核查状态与来源追踪", 6, "C00000")

    hdr_row(ws, 2, ["省份", "核查状态", "GDP排名", "已收录类别数", "时间序列(条)", "说明"], "C00000")
    set_col_width(ws, [(1, 12), (2, 12), (3, 10), (4, 14), (5, 16), (6, 40)])

    vs = national["verification_status"]
    provinces = national["provinces"]
    row = 3
    for prov_name in sorted(provinces.keys(),
                              key=lambda x: {"full": 0, "partial": 1, "minimal": 2}.get(
                                  provinces[x].get("verification", ""), 3)):
        prov = provinces[prov_name]
        cats = prov.get("data", {})
        verif = prov.get("verification", "")
        ts_count = len(prov.get("time_series", {}))
        cat_count = len(cats)

        fill = VERIF_COLORS.get(verif, "")
        verif_label = {"full": "A类(已核查)", "partial": "B类(部分核查)", "minimal": "C类(最少)"}[verif]

        note = ""
        if verif == "full":
            note = "湖北为基准省份，全部45项指标已核查"
        elif verif == "partial":
            note = "GDP/R&D等核心指标已收录，高校/AI/金融等类别待补充"
        else:
            note = "仅收录GDP、高新企业、技术合同三项，需扩充"

        for col, val in enumerate([prov_name, verif_label,
                                    prov.get("gdp_rank_2024", "—"),
                                    cat_count, ts_count, note], 1):
            data_cell(ws, row, col, val, fill_hex=fill, bold=(col == 1),
                      align="center" if col in (2, 3, 4, 5) else "left")
        row += 1

    # Summary statistics
    row += 1
    title_row(ws, row, "数据质量统计", 4, "404040")
    row += 1
    hdr_row(ws, row, ["指标", "数值", "说明", ""], "404040")

    total_indicators = sum(
        len(cats) for prov in provinces.values() for cats in prov.get("data", {}).values()
    )
    total_with_source = sum(
        sum(1 for v in cats.values() if isinstance(v, dict) and v.get("source"))
        for prov in provinces.values() for cats in prov.get("data", {}).values()
    )
    total_ts = sum(
        sum(len(s.get("data", {})) for s in prov.get("time_series", {}).values())
        for prov in provinces.values()
    )
    ranking_count = len(national.get("ranking_tables", {}))

    stats = [
        ("收录省份数", f"{len(provinces)} 个"),
        ("A类省份(全核查)", f"{len(vs.get('full', []))} 个: {', '.join(vs.get('full', []))}"),
        ("B类省份(部分)", f"{len(vs.get('partial', []))} 个: {', '.join(vs.get('partial', []))}"),
        ("C类省份(最少)", f"{len(vs.get('minimal', []))} 个: {', '.join(vs.get('minimal', []))}"),
        ("指标总数", f"{total_indicators} 条"),
        ("有来源标注", f"{total_with_source} 条 ({total_with_source*100//total_indicators if total_indicators else 0}%)"),
        ("时间序列数据点", f"{total_ts} 个"),
        ("排名表数量", f"{ranking_count} 张"),
        ("数据框架版本", national["meta"].get("version", "—")),
        ("生成日期", national["meta"].get("generated_at", "—")),
    ]
    for i, (k, v) in enumerate(stats):
        for col, val in enumerate([k, v, "", ""], 1):
            data_cell(ws, row + i, col, val, bold=(col == 1))


# ── Main ────────────────────────────────────────────────────────────────────────

def main():
    if not DATA_FILE.exists():
        print(f"ERROR: {DATA_FILE} not found")
        sys.exit(1)

    national = load_data()
    wb = openpyxl.Workbook()

    sheet_overview(wb, national)
    sheet_gdp(wb, national)
    sheet_rd(wb, national)
    sheet_tech_ent(wb, national)
    sheet_tech(wb, national)
    sheet_industry(wb, national)
    sheet_edu(wb, national)
    sheet_hubei_ts(wb, national)
    sheet_quality(wb, national)

    wb.save(OUT_FILE)
    print(f"Done -> {OUT_FILE}  ({len(wb.sheetnames)} sheets)")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()

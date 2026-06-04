#!/usr/bin/env python3
"""导出湖北省科技创新数据Excel（含数据分类标注）"""

import subprocess

subprocess.run(['pip', 'install', 'openpyxl', 'pandas', '-q'], capture_output=True)

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── 路径配置（基于脚本位置自动计算）───────────────────────────────
_PROJECT_ROOT = Path(__file__).parent.parent
_DATA_DIR = _PROJECT_ROOT / "data"
_DATA_FILE = _DATA_DIR / "hubei_tech_data_2026.json"
_OUT_FILE = _DATA_DIR / "湖北省科技创新数据_2026.xlsx"

with open(_DATA_FILE, encoding='utf-8') as f:
    data = json.load(f)

wb = Workbook()

# ===== 样式 =====
HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL_BLUE = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FILL_GREEN = PatternFill(start_color='1D6B3C', end_color='1D6B3C', fill_type='solid')
HEADER_FILL_ORANGE = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid')
HEADER_FILL_RED = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
A_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
B_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
C_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

def hdr(cell, text, fill=None):
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = fill or HEADER_FILL_BLUE
    cell.alignment = CENTER
    cell.border = THIN_BORDER

def body(cell, text='', bold=False):
    if text != '':
        cell.value = text
    cell.font = Font(name='微软雅黑', size=10, bold=bold)
    cell.alignment = LEFT
    cell.border = THIN_BORDER

# ═══════════════════════════════════════════════════════════════
# Sheet 1: 完整数据清单（含分类）
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '完整数据清单'
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 26
ws1.column_dimensions['D'].width = 16
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 45
ws1.column_dimensions['G'].width = 12
ws1.column_dimensions['H'].width = 50
ws1.column_dimensions['I'].width = 20

for c, h in enumerate(['序号','数据分类','指标','数值','单位','备注','年份','数据来源','MCP/手动录入说明'], 1):
    hdr(ws1.cell(row=1, column=c), h)

# A类真实数据
row = 2
for idx, item in enumerate(data['real_data']['data'], 1):
    row_fill = A_FILL
    cls_label = 'A'
    mcp_note = '可直接引用，无需补充'
    body(ws1.cell(row=row, column=1), idx)
    ws1.cell(row=row, column=1).fill = row_fill
    ws1.cell(row=row, column=1).alignment = CENTER
    ws1.cell(row=row, column=2).fill = row_fill
    body(ws1.cell(row=row, column=2), 'A 真实数据', bold=True)
    ws1.cell(row=row, column=3).fill = row_fill
    body(ws1.cell(row=row, column=3), item['indicator'])
    cell4 = ws1.cell(row=row, column=4)
    cell4.fill = row_fill
    val = item['value']
    if isinstance(val, (int, float)):
        cell4.value = float(val)
        cell4.number_format = '#,##0.##'
    else:
        cell4.value = val
    body(cell4)
    cell4.alignment = CENTER
    ws1.cell(row=row, column=5).fill = row_fill
    body(ws1.cell(row=row, column=5), item['unit'])
    ws1.cell(row=row, column=5).alignment = CENTER
    ws1.cell(row=row, column=6).fill = row_fill
    body(ws1.cell(row=row, column=6), item.get('note', ''))
    ws1.cell(row=row, column=7).fill = row_fill
    body(ws1.cell(row=row, column=7), item['year'])
    ws1.cell(row=row, column=7).alignment = CENTER
    ws1.cell(row=row, column=8).fill = row_fill
    body(ws1.cell(row=row, column=8), item['source'])
    ws1.cell(row=row, column=9).fill = row_fill
    body(ws1.cell(row=row, column=9), mcp_note)
    for c in range(1, 10):
        ws1.cell(row=row, column=c).border = THIN_BORDER
    row += 1

# B类MCP可补充
a_count = len(data['real_data']['data'])
for idx, item in enumerate(data['mcp_available']['indicators'], 1):
    row_fill = B_FILL
    body(ws1.cell(row=row, column=1), a_count + idx)
    ws1.cell(row=row, column=1).fill = row_fill
    ws1.cell(row=row, column=1).alignment = CENTER
    ws1.cell(row=row, column=2).fill = row_fill
    body(ws1.cell(row=row, column=2), 'B MCP可补充', bold=True)
    ws1.cell(row=row, column=3).fill = row_fill
    body(ws1.cell(row=row, column=3), item['indicator'])
    ws1.cell(row=row, column=4).fill = row_fill
    body(ws1.cell(row=row, column=4), item.get('gap', ''))
    ws1.cell(row=row, column=5).fill = row_fill
    body(ws1.cell(row=row, column=5), '-')
    ws1.cell(row=row, column=5).alignment = CENTER
    ws1.cell(row=row, column=6).fill = row_fill
    body(ws1.cell(row=row, column=6), item.get('mcp_source', ''))
    ws1.cell(row=row, column=7).fill = row_fill
    body(ws1.cell(row=row, column=7), '-')
    ws1.cell(row=row, column=7).alignment = CENTER
    ws1.cell(row=row, column=8).fill = row_fill
    body(ws1.cell(row=row, column=8), item.get('source_url', ''))
    ws1.cell(row=row, column=9).fill = row_fill
    body(ws1.cell(row=row, column=9), f"优先级: {item['mcp_priority']} | 状态: {item['mcp_status']}")
    for c in range(1, 10):
        ws1.cell(row=row, column=c).border = THIN_BORDER
    row += 1

# C类手动录入
b_count = len(data['mcp_available']['indicators'])
for idx, item in enumerate(data['manual_entry']['indicators'], 1):
    row_fill = C_FILL
    body(ws1.cell(row=row, column=1), a_count + b_count + idx)
    ws1.cell(row=row, column=1).fill = row_fill
    ws1.cell(row=row, column=1).alignment = CENTER
    ws1.cell(row=row, column=2).fill = row_fill
    body(ws1.cell(row=row, column=2), 'C 手动录入', bold=True)
    ws1.cell(row=row, column=3).fill = row_fill
    body(ws1.cell(row=row, column=3), item['indicator'])
    ws1.cell(row=row, column=4).fill = row_fill
    body(ws1.cell(row=row, column=4), item.get('value', ''))
    ws1.cell(row=row, column=5).fill = row_fill
    body(ws1.cell(row=row, column=5), '-')
    ws1.cell(row=row, column=5).alignment = CENTER
    ws1.cell(row=row, column=6).fill = row_fill
    body(ws1.cell(row=row, column=6), item.get('note', ''))
    ws1.cell(row=row, column=7).fill = row_fill
    body(ws1.cell(row=row, column=7), item.get('year', ''))
    ws1.cell(row=row, column=7).alignment = CENTER
    ws1.cell(row=row, column=8).fill = row_fill
    body(ws1.cell(row=row, column=8), item['source'])
    ws1.cell(row=row, column=9).fill = row_fill
    body(ws1.cell(row=row, column=9), f"更新: {item.get('update_frequency', item.get('note', ''))}")
    for c in range(1, 10):
        ws1.cell(row=row, column=c).border = THIN_BORDER
    row += 1

ws1.freeze_panes = 'A2'

# ═══════════════════════════════════════════════════════════════
# Sheet 2: 历年数据序列
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('历年数据序列')
ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 28
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 12

years = ['2005','2007','2010','2015','2016','2019','2020','2021','2022','2023','2024','2025']
for i, y in enumerate(years):
    ws2.column_dimensions[get_column_letter(i+5)].width = 10

hdr(ws2.cell(row=1,column=1), '序号')
hdr(ws2.cell(row=1,column=2), '指标名称')
hdr(ws2.cell(row=1,column=3), '单位')
hdr(ws2.cell(row=1,column=4), '数据类型')
for i, y in enumerate(years):
    hdr(ws2.cell(row=1,column=i+5), f'{y}年')

ts = data['real_data']['time_series']
row = 2
idx = 1
for series_name, series_data in ts.items():
    dt = series_data.get('data_type', 'A')
    fill = A_FILL
    for c in range(1, 5):
        ws2.cell(row=row, column=c).fill = fill
    body(ws2.cell(row=row,column=1), idx)
    ws2.cell(row=row,column=1).alignment = CENTER
    body(ws2.cell(row=row,column=2), series_name)
    body(ws2.cell(row=row,column=3), series_data['unit'])
    body(ws2.cell(row=row,column=4), 'A 真实数据')
    ws2.cell(row=row,column=4).alignment = CENTER
    for i, y in enumerate(years):
        cell = ws2.cell(row=row, column=i+5)
        val = series_data['data'].get(y)
        if val is not None:
            cell.value = float(val)
            cell.number_format = '#,##0.##'
            body(cell)
        else:
            cell.value = '-'
            body(cell)
            cell.font = Font(name='微软雅黑', size=10, color='C00000')
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    row += 1
    idx += 1

ws2.freeze_panes = 'E2'

# ═══════════════════════════════════════════════════════════════
# Sheet 3: MCP开发计划
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('MCP开发计划')
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 28
ws3.column_dimensions['C'].width = 40
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 15
ws3.column_dimensions['F'].width = 15

for c, h in enumerate(['MCP服务名','描述','数据源','工具列表','优先级','开发难度'], 1):
    hdr(ws3.cell(row=1,column=c), h)

mcp_servers = data.get('mcp_services', {}).get('servers', data.get('mcp_development_plan', {}).get('servers', []))
row = 2
for srv in mcp_servers:
    body(ws3.cell(row=row,column=1), srv['name'])
    body(ws3.cell(row=row,column=2), srv.get('description', srv.get('description', '')))
    # data_sources 可能是新格式的工具数，也可能是旧格式的列表
    tools_list = srv.get('tool_list', srv.get('tools', []))
    body(ws3.cell(row=row,column=3), f"{len(tools_list)} 个工具")
    body(ws3.cell(row=row,column=4), ' | '.join(tools_list[:8]) if isinstance(tools_list, list) and tools_list else str(tools_list))
    # 优先级和开发难度（可能是新格式中的标签）
    prio = srv.get('priority', srv.get('status', '已部署'))
    cell5 = ws3.cell(row=row,column=5)
    cell5.value = prio
    cell5.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    cell5.alignment = CENTER
    cell5.border = THIN_BORDER
    cell5.fill = PatternFill(start_color='1D6B3C', end_color='1D6B3C', fill_type='solid')
    eff = srv.get('development_effort', '已部署')
    cell6 = ws3.cell(row=row,column=6)
    cell6.value = eff
    cell6.font = Font(name='微软雅黑', size=10)
    cell6.alignment = CENTER
    cell6.border = THIN_BORDER
    cell6.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    row += 1

ws3.freeze_panes = 'A2'

# ═══════════════════════════════════════════════════════════════
# Sheet 4: 手动录入清单
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('手动录入清单')
ws4.column_dimensions['A'].width = 6
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 28
ws4.column_dimensions['D'].width = 35
ws4.column_dimensions['E'].width = 12
ws4.column_dimensions['F'].width = 30
ws4.column_dimensions['G'].width = 22

for c, h in enumerate(['序号','类别','指标','当前值','年份','数据来源','更新频率'], 1):
    hdr(ws4.cell(row=1,column=c), h)

row = 2
for idx, item in enumerate(data['manual_entry']['indicators'], 1):
    for c in range(1,8):
        ws4.cell(row=row,column=c).fill = C_FILL
    body(ws4.cell(row=row,column=1), idx)
    ws4.cell(row=row,column=1).alignment = CENTER
    body(ws4.cell(row=row,column=2), item['category'])
    body(ws4.cell(row=row,column=3), item['indicator'])
    body(ws4.cell(row=row,column=4), item.get('value',''))
    body(ws4.cell(row=row,column=5), item.get('year',''))
    ws4.cell(row=row,column=5).alignment = CENTER
    body(ws4.cell(row=row,column=6), item['source'])
    body(ws4.cell(row=row,column=7), item.get('update_frequency', item.get('note', '')))
    for c in range(1,8):
        ws4.cell(row=row,column=c).border = THIN_BORDER
    row += 1

ws4.freeze_panes = 'A2'

# ═══════════════════════════════════════════════════════════════
# 保存
# ═══════════════════════════════════════════════════════════════
out = _OUT_FILE
wb.save(out)
print(f'Excel已保存: {out}')
print(f'工作表: {wb.sheetnames}')
